import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

diretorio = "C:/relatoriosEstoque"

diretorio_consolidacoes = os.path.join(diretorio, "consolidacoes")
os.makedirs(diretorio_consolidacoes, exist_ok=True)  # Criar a pasta se não existir

dados_unificados = []

for arquivo in os.listdir(diretorio):
    if arquivo.endswith(".xlsx") or arquivo.endswith(".xls"):  
        caminho_arquivo = os.path.join(diretorio, arquivo)
        
       
        wb = load_workbook(caminho_arquivo, data_only=True)
        sheet = wb.active 
        
        id_filial = sheet["D2"].value
        
        dados = []
        for linha in sheet.iter_rows(min_row=29, max_row=34, min_col=1, max_col=6, values_only=True):
            if linha[0] != "Material":  # Evitar repetição de cabeçalho
                dados.append(linha)
        
        df_temp = pd.DataFrame(dados, columns=["Material", "Est Ant", "Entradas", "Est. Fis", "Saídas", "Cons/m3"])
        df_temp.insert(0, "Filial", id_filial)  # Inserir a coluna da filial
       
        dados_unificados.append(df_temp)


if dados_unificados:
    df_final = pd.concat(dados_unificados, ignore_index=True)
    #p nome
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nome_arquivo = os.path.join(diretorio_consolidacoes, f"planilha_unificada_{timestamp}.xlsx")
    
    
    df_final.to_excel(nome_arquivo, index=False)
    print(f"Planilha unificada criada com sucesso: {nome_arquivo}")
else:
    print("Nenhum dado encontrado para consolidar.")
